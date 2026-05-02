import os
import re
import pdfplumber
import docx
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook #for cell fromating
from openpyxl.styles import Alignment #for wrapping info in each cell
 
# Regex patterns for email + phone
EMAIL_REGEX = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
PHONE_REGEX = r"(\+?\d{1,2}[-.\s]?)?(\(?\d{3}\)?[-.\s]?)?\d{3}[-.\s]?\d{4}"
 
# Load or create processed log
def load_processed_log(log_path="processed_log.txt"):
    if not os.path.exists(log_path):
        return set()
    with open(log_path, "r") as f:
        return set(line.strip() for line in f.readlines())
 
def update_processed_log(filename, log_path="processed_log.txt"):
    with open(log_path, "a") as f:
        f.write(filename + "\n")
 
 
# Extract text from PDF or DOCX
def extract_text_from_pdf(filepath):
    try:
        with pdfplumber.open(filepath) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text(x_tolerance=1, y_tolerance=1) or ""
            return text
    except Exception as e:
        print(f"Error reading PDF {filepath}: {e}")
        return ""
        
def extract_text_from_docx(filepath):
    try:
        doc = docx.Document(filepath)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        print(f"Error reading DOCX {filepath}: {e}")
        return ""
 
 
# Clean text
def clean_text(text):
    # Keep multiple spaces, but remove weird characters and normalize newlines
    text = text.replace("\t", " ") # convert tabs to spaces
    text = text.replace("\r", "")  # remove carriage returns
    text = re.sub(r"[￣]+", "", text)  # remove decorative line characters (￣ used as dividers)
    text = re.sub(r"[ ]{2,}", "  ", text) # keep double spaces, collapse 3+ to 2
    text = re.sub(r"\n{2,}", "\n", text)  # collapse multiple blank lines
    return text.strip()
 
 
# Extract fields (name, email, phone, skills, etc.)
def extract_email(text):
    match = re.search(EMAIL_REGEX, text)
    return match.group(0) if match else ""
 
def extract_phone(text):
    match = re.search(PHONE_REGEX, text)
    return match.group(0) if match else ""
 
def extract_name(text):
    # Heuristic: first line with 2–3 words, capitalized
    lines = text.split("\n")
    for line in lines[:5]:
        if 2 <= len(line.split()) <= 4 and line[0].isupper():
            return line.strip()
    return ""
 
 
SECTION_HEADERS = [
    "research experience", "work experience", "work history",  "professional experience",
    "project experience", "other education", "additional education", "education & skills",
    "technical skills", "computational skills", "computer skills", "education",
    "experience", "employment", "skills", "projects", "certifications", "publications",
    "summary", "activities", "extracurricular", "honors", "awards",
]
 
# Pre-compile a regex that matches any header only at the start of a line
_HEADER_PATTERN = re.compile(
    r"(?:^|\n)(" + "|".join(re.escape(h) for h in SECTION_HEADERS) + r")\b",
    re.IGNORECASE
)
 
def _find_section_start(text_lower, keyword):
    """Return the index where keyword begins as a standalone section header line."""
    pattern = re.compile(r"(?:^|\n)" + re.escape(keyword) + r"\b", re.IGNORECASE)
    m = pattern.search(text_lower)
    if m:
        return m.start() + (1 if text_lower[m.start()] == "\n" else 0)
    return -1
 
def _extract_one(text, kw):
    """Extract a single section matching kw at start of line. Returns (start, content) or None."""
    text_lower = text.lower()
    start = _find_section_start(text_lower, kw)
    if start == -1:
        return None
    next_section_start = len(text)
    search_from = start + len(kw)
    for m in _HEADER_PATTERN.finditer(text, search_from):
        header_start = m.start() + (1 if text[m.start()] == "\n" else 0)
        if header_start < next_section_start:
            next_section_start = header_start
            break
    return (start, text[start:next_section_start].strip())
 
def extract_section(text, keywords):
    """Return the first matching section. For experience, see extract_all_sections."""
    for kw in keywords:
        result = _extract_one(text, kw)
        if result:
            return result[1]
    return ""
 
def extract_all_sections(text, keywords):
    """Find ALL matching sections (e.g. Research Experience + Experience) and join them."""
    found = []
    for kw in keywords:
        result = _extract_one(text, kw)
        if result:
            found.append(result)
    if not found:
        return ""
    # Sort by position in document, deduplicate overlapping sections
    found.sort(key=lambda x: x[0])
    seen_starts = set()
    parts = []
    for start, content in found:
        if start not in seen_starts:
            seen_starts.add(start)
            parts.append(content)
    return "\n\n".join(parts)
 
def parse_resume(text):
    return {
        "Name": extract_name(text),
        "Email": extract_email(text),
        "Phone": extract_phone(text),
        "Education": extract_section(text, ["education"]),
        "Skills": extract_section(text, ["computational skills", "computer skills", "technical skills", "skills"]),
        "Experience": extract_all_sections(text, ["research experience", "work experience", "professional experience", "experience", "work history", "employment"]),
        "Projects": extract_section(text, ["projects", "project experience", "project"]),
        "Comments": ""
    }
 
 
 
def format_fixed_cell_sizes(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
 
    ws.auto_filter.ref = ws.dimensions
 
 
    # Set column width for ALL columns
    for column in ws.columns:
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = 60
 
    # Set row height for ALL rows EXCEPT the first row
    for row in ws.iter_rows(min_row=2):  # start at row 2
        row_number = row[0].row
        ws.row_dimensions[row_number].height = 350
 
    # Wrap text in ALL cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
 
    wb.save(excel_path)
 
# Function that processes the resumes with the other functions
def process_resumes(resume_folder="Resumes", output_file="output.xlsx"):
    resume_folder = Path(resume_folder)
    processed = load_processed_log()
    # Normalize log entries to lowercase+stripped for reliable comparison
    processed_normalized = {p.strip().lower() for p in processed}
 
    rows = []
 
    for file in resume_folder.iterdir():
        if file.name.strip().lower() in processed_normalized:
            print(f"Already processed, skipping: {file.name}")
            continue
 
        if file.suffix.lower() == ".pdf":
            raw_text = extract_text_from_pdf(file)
        elif file.suffix.lower() in [".docx", ".doc"]:
            raw_text = extract_text_from_docx(file)
        else:
            print(f"Skipping unsupported file: {file.name}")
            continue
 
        cleaned = clean_text(raw_text)
        data = parse_resume(cleaned)
        data["Filename"] = file.name
 
        rows.append(data)
        # NOTE: log updated after successful Excel write below
 
    if rows:
        new_df = pd.DataFrame(rows)
 
        # Append to existing Excel if it exists, otherwise create fresh
        if os.path.exists(output_file):
            existing_df = pd.read_excel(output_file)
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            combined_df = new_df
 
        combined_df.to_excel(output_file, index=False)
 
        # Fix the column sizes and the alignment
        format_fixed_cell_sizes(output_file)
 
        # Only mark as processed AFTER Excel is successfully written
        for row in rows:
            update_processed_log(row["Filename"])
 
        print(f"Excel updated: {output_file} — added {len(rows)} resume(s)")
    else:
        print("No new resumes to process.")
 
 
# Runs the script
if __name__ == "__main__":
    process_resumes()